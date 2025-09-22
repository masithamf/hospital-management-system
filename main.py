from fastapi import FastAPI, Depends, HTTPException, status, Request, Form, Response, Cookie
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from datetime import datetime, timedelta, date
from typing import Optional
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.utils import get_column_letter
import models, schemas, auth
from database import engine, get_db

models.Base.metadata.create_all(bind=engine)

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Auth routes
@app.post("/login")
async def login(
    request: Request,
    response: Response,
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=30)
    access_token = auth.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )

    # # Mengembalikan token sebagai JSON
    # return {"access_token": access_token, "token_type": "bearer"}

    
    # Ambil data untuk dashboard
    patients = db.query(models.Patient).all()
    total_patients = len(patients)
    
    # Langsung render dashboard dengan cookie yang sudah diset
    response = templates.TemplateResponse("dashboard.html", {
        "request": request,
        "total_patients": total_patients,
        "patients": patients,
        "user": user
    })
    
    # Set cookie pada response
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        max_age=1800,
        samesite="lax",
        secure=False,
        path="/"
    )
    
    return response

@app.post("/logout")
async def logout(response: Response):
    response.delete_cookie("access_token")
    return RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


# Patient routes
@app.get("/patients", response_class=HTMLResponse)
async def list_patients(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    patients = db.query(models.Patient).all()
    return templates.TemplateResponse(
        "patients/list.html",
        {"request": request, "patients": patients, "user": current_user}
    )

@app.get("/patients/create", response_class=HTMLResponse)
async def create_patient_form(
    request: Request,
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    return templates.TemplateResponse("patients/create.html", {"request": request, "user": current_user})

@app.post("/patients")
async def create_patient(
    nama: str = Form(...),
    tanggal_lahir: str = Form(...),
    diagnosis: str = Form(...),
    tindakan: str = Form(...),
    dokter: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    patient = models.Patient(
        nama=nama,
        tanggal_lahir=datetime.strptime(tanggal_lahir, "%Y-%m-%d").date(),
        diagnosis=diagnosis,
        tindakan=tindakan,
        dokter=dokter
    )
    db.add(patient)
    db.commit()
    db.refresh(patient)
    return RedirectResponse(url="/patients", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/patients/{patient_id}/edit", response_class=HTMLResponse)
async def edit_patient_form(
    request: Request,
    patient_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    patient = db.query(models.Patient).filter(models.Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return templates.TemplateResponse(
        "patients/edit.html",
        {"request": request, "patient": patient, "user": current_user}
    )

@app.post("/patients/{patient_id}")
async def update_patient(
    patient_id: int,
    nama: str = Form(...),
    tanggal_lahir: str = Form(...),
    diagnosis: str = Form(...),
    tindakan: str = Form(...),
    dokter: str = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    patient = db.query(models.Patient).filter(models.Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    patient.nama = nama
    patient.tanggal_lahir = datetime.strptime(tanggal_lahir, "%Y-%m-%d").date()
    patient.diagnosis = diagnosis
    patient.tindakan = tindakan
    patient.dokter = dokter
    
    db.commit()
    return RedirectResponse(url="/patients", status_code=status.HTTP_303_SEE_OTHER)

@app.post("/patients/{patient_id}/delete")
async def delete_patient(
    patient_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    patient = db.query(models.Patient).filter(models.Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    db.delete(patient)
    db.commit()
    return RedirectResponse(url="/patients", status_code=status.HTTP_303_SEE_OTHER)

# Dashboard routes
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    query = db.query(models.Patient)
    
    if from_date and to_date:
        query = query.filter(
            models.Patient.tanggal_kunjungan.between(from_date, to_date)
        )
    
    if search:
        query = query.filter(models.Patient.nama.ilike(f"%{search}%"))
    
    patients = query.all()
    total_patients = len(patients)
    
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "total_patients": total_patients,
            "patients": patients,
            "user": current_user
        }
    )

# Export route
@app.get("/export")
async def export_patients(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    patients = db.query(models.Patient).all()
    
    # Create Excel workbook using openpyxl directly
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    
    # Headers
    headers = ["ID", "Nama", "Tanggal Lahir", "Tanggal Kunjungan", "Diagnosis", "Tindakan", "Dokter"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data rows
    for row, patient in enumerate(patients, 2):
        ws.cell(row=row, column=1, value=patient.id)
        ws.cell(row=row, column=2, value=patient.nama or "")
        ws.cell(row=row, column=3, value=str(patient.tanggal_lahir) if patient.tanggal_lahir else "")
        ws.cell(row=row, column=4, value=str(patient.tanggal_kunjungan) if patient.tanggal_kunjungan else "")
        ws.cell(row=row, column=5, value=patient.diagnosis or "")
        ws.cell(row=row, column=6, value=patient.tindakan or "")
        ws.cell(row=row, column=7, value=patient.dokter or "")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="patients.xlsx"'
    }
    
    return StreamingResponse(
        output,
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Import route - Accept JSON data
@app.post("/import")
async def import_patients(
    patients_data: list[dict],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    """
    Import patients from external source via POST JSON
    Expected format:
    [
        {
            "nama": "Patient Name",
            "tanggal_kunjungan": "2024-01-20T10:30:00" (optional, defaults to now)
        }
    ]
    """
    imported_count = 0
    
    for patient_data in patients_data:
        # Validate required fields
        if not patient_data.get("nama"):
            continue
            
        # Set default values for missing fields
        patient_info = {
            "nama": patient_data["nama"],
            "tanggal_lahir": patient_data.get("tanggal_lahir", datetime.now().date()),
            "diagnosis": patient_data.get("diagnosis", "General Checkup"),
            "tindakan": patient_data.get("tindakan", "Consultation"),
            "dokter": patient_data.get("dokter", current_user.username)
        }
        
        # Handle tanggal_kunjungan if provided
        if "tanggal_kunjungan" in patient_data:
            try:
                # Parse datetime string if provided
                if isinstance(patient_data["tanggal_kunjungan"], str):
                    patient_info["tanggal_kunjungan"] = datetime.fromisoformat(patient_data["tanggal_kunjungan"])
                else:
                    patient_info["tanggal_kunjungan"] = patient_data["tanggal_kunjungan"]
            except:
                # Use default if parsing fails
                pass
        
        patient = models.Patient(**patient_info)
        db.add(patient)
        imported_count += 1
    
    db.commit()
    
    # Redirect ke patients setelah import berhasil untuk mencegah duplicate import saat refresh
    return RedirectResponse(
        url=f"/patients?imported={imported_count}", 
        status_code=status.HTTP_303_SEE_OTHER
    )

# Import dummy data route (for testing)
@app.post("/import/dummy")
async def import_dummy_patients(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_doctor)
):
    """Import dummy patients for testing purposes"""
    dummy_data = [
        {
            "nama": "John Doe",
            "tanggal_lahir": date(1990, 5, 15),
            "tanggal_kunjungan": datetime.now(),
            "diagnosis": "Flu",
            "tindakan": "Medication and rest",
            "dokter": "Dr. Smith"
        },
        {
            "nama": "Jane Smith",
            "tanggal_lahir": date(1985, 8, 22),
            "tanggal_kunjungan": datetime.now(),
            "diagnosis": "Fever",
            "tindakan": "Paracetamol and monitoring",
            "dokter": "Dr. Johnson"
        },
        {
            "nama": "Bob Wilson",
            "tanggal_lahir": date(1978, 12, 3),
            "tanggal_kunjungan": datetime.now(),
            "diagnosis": "Hypertension",
            "tindakan": "Blood pressure medication",
            "dokter": "Dr. Brown"
        }
    ]
    
    for data in dummy_data:
        patient = models.Patient(**data)
        db.add(patient)
    
    db.commit()
    return RedirectResponse(url="/patients", status_code=status.HTTP_303_SEE_OTHER)
